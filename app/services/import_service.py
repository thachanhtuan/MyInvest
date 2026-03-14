from __future__ import annotations

import datetime
import logging
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import IMPORT_LOG_FILE, SHEET_NAMES
from app.models import Account, Asset, AssetValuation, BondInfo, Transaction
from app.schemas import ImportResult
from app.utils.logging_util import flush_all_handlers, setup_file_handler

logger = logging.getLogger(__name__)
setup_file_handler(logger, IMPORT_LOG_FILE)  # Configure once at import time


def _cell_to_date(value: object) -> datetime.date | None:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


def _cell_to_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _cell_to_int(value: object) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _cell_to_bool(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "да", "yes")


def _cell_to_str(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _read_sheet_rows(ws) -> list[dict[str, object]]:
    """Read worksheet rows as list of dicts keyed by header names."""
    headers: list[str | None] = []
    for cell in ws[1]:
        headers.append(cell.value)

    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_data: dict[str, object] = {}
        all_none = True
        for idx, cell in enumerate(row):
            if idx < len(headers) and headers[idx] is not None:
                row_data[headers[idx]] = cell.value
                if cell.value is not None:
                    all_none = False
        if not all_none:
            rows.append(row_data)
    return rows


def _import_accounts(db: Session, rows: list[dict]) -> int:
    count = 0
    for row in rows:
        acc_id = _cell_to_str(row.get("id"))
        if not acc_id:
            continue
        account = Account(
            id=acc_id,
            name=_cell_to_str(row.get("name")) or acc_id,
            type=_cell_to_str(row.get("type")) or "broker",
            broker_or_bank=_cell_to_str(row.get("broker_or_bank")) or "",
            currency=_cell_to_str(row.get("currency")) or "RUB",
            notes=_cell_to_str(row.get("notes")),
        )
        db.add(account)
        count += 1
    return count


def _import_assets(db: Session, rows: list[dict]) -> int:
    count = 0
    for row in rows:
        ticker = _cell_to_str(row.get("ticker"))
        if not ticker:
            continue
        asset = Asset(
            ticker=ticker,
            name=_cell_to_str(row.get("name")) or ticker,
            asset_class=_cell_to_str(row.get("asset_class")) or "other",
            currency=_cell_to_str(row.get("currency")) or "RUB",
            isin=_cell_to_str(row.get("isin")),
            notes=_cell_to_str(row.get("notes")),
            target_min=_cell_to_float(row.get("target_min")),
            target_max=_cell_to_float(row.get("target_max")),
            target_pct=_cell_to_float(row.get("target_pct")),
            target_date=_cell_to_date(row.get("target_date")),
            exchange=_cell_to_str(row.get("exchange")),
        )
        db.add(asset)
        count += 1
    return count


def _import_bonds(db: Session, rows: list[dict]) -> int:
    count = 0
    for row in rows:
        ticker = _cell_to_str(row.get("ticker"))
        if not ticker:
            continue
        bond = BondInfo(
            ticker=ticker,
            face_value=_cell_to_float(row.get("face_value")) or 1000,
            base_currency=_cell_to_str(row.get("base_currency")),
            coupon_rate=_cell_to_float(row.get("coupon_rate")),
            coupon_sum=_cell_to_float(row.get("coupon_sum")),
            coupon_currency=_cell_to_str(row.get("coupon_currency")),
            coupon_frequency_year=_cell_to_int(row.get("coupon_frequency_year")),
            coupon_frequency_day=_cell_to_int(row.get("coupon_frequency_day")) or 30,
            maturity_date=_cell_to_date(row.get("maturity_date")),
            offer_date=_cell_to_date(row.get("offer_date")),
            first_coupon_date=_cell_to_date(row.get("first_coupon_date")),
            is_amortizing=_cell_to_bool(row.get("is_amortizing")),
        )
        db.add(bond)
        count += 1
    return count


def _compute_amount(row: dict) -> float | None:
    """Compute amount per EXCEL_FORMAT.md rules."""
    amount = _cell_to_float(row.get("amount"))
    if amount is not None:
        return amount
    quantity = _cell_to_float(row.get("quantity"))
    price = _cell_to_float(row.get("price"))
    if quantity is not None and price is not None:
        return quantity * price
    return None


def _compute_total_amount(row: dict, amount: float | None) -> float | None:
    """Compute total_amount per EXCEL_FORMAT.md rules."""
    total = _cell_to_float(row.get("total_amount"))
    if total is not None:
        return total
    if amount is None:
        return None
    commission = _cell_to_float(row.get("commission")) or 0.0
    nkd = _cell_to_float(row.get("nkd")) or 0.0
    tx_type = _cell_to_str(row.get("tx_type")) or ""
    if tx_type == "sell":
        return amount - commission
    return amount + commission + nkd


def _import_transactions(db: Session, rows: list[dict]) -> int:
    count = 0
    for row in rows:
        tx_date = _cell_to_date(row.get("date"))
        account_id = _cell_to_str(row.get("account_id"))
        tx_type = _cell_to_str(row.get("tx_type"))
        if not tx_date or not account_id or not tx_type:
            continue

        amount = _compute_amount(row)
        total_amount = _compute_total_amount(row, amount)

        tx = Transaction(
            date=tx_date,
            account_id=account_id,
            ticker=_cell_to_str(row.get("ticker")),
            tx_type=tx_type,
            quantity=_cell_to_float(row.get("quantity")),
            price=_cell_to_float(row.get("price")),
            amount=amount,
            nominal_amount=_cell_to_float(row.get("nominal_amount")),
            nominal_currency=_cell_to_str(row.get("nominal_currency")),
            nkd=_cell_to_float(row.get("nkd")),
            commission=_cell_to_float(row.get("commission")),
            total_amount=total_amount,
            currency=_cell_to_str(row.get("currency")) or "RUB",
            broker=_cell_to_str(row.get("broker")),
            notes=_cell_to_str(row.get("notes")),
        )
        db.add(tx)
        count += 1
    return count


def _import_valuations(db: Session, rows: list[dict]) -> int:
    count = 0
    for row in rows:
        val_date = _cell_to_date(row.get("date"))
        ticker = _cell_to_str(row.get("ticker"))
        if not val_date or not ticker:
            continue
        valuation = AssetValuation(
            date=val_date,
            ticker=ticker,
            value=_cell_to_float(row.get("value")),
            currency=_cell_to_str(row.get("currency")),
            nominal_value=_cell_to_float(row.get("nominal_value")),
            nominal_currency=_cell_to_str(row.get("nominal_currency")),
            value_pct=_cell_to_float(row.get("value_pct")),
        )
        db.add(valuation)
        count += 1
    return count


def import_excel(db: Session, file_path: Path) -> ImportResult:
    """Import data from Excel file into the database.

    Strategy: clear all tables and re-insert.
    Order: Accounts -> Assets -> BondInfo -> Transactions -> AssetValuations
    """
    result = ImportResult()

    logger.info("=" * 80)
    logger.info(f"Starting import from: {file_path}")

    try:
        wb = load_workbook(file_path, data_only=True)
        logger.info(f"Excel file loaded, sheets: {wb.sheetnames}")
    except Exception as e:
        logger.error(f"Failed to open file: {e}")
        result.errors.append(f"Не удалось открыть файл: {e}")
        flush_all_handlers(logger)
        return result

    sheet_names_in_file = wb.sheetnames

    # Clear tables in reverse FK order
    logger.info("Clearing existing data...")
    db.query(AssetValuation).delete()
    db.query(Transaction).delete()
    db.query(BondInfo).delete()
    db.query(Asset).delete()
    db.query(Account).delete()
    db.flush()
    logger.info("Data cleared")

    # Import accounts
    sheet_name = SHEET_NAMES["accounts"]
    if sheet_name in sheet_names_in_file:
        logger.info(f"Importing accounts from «{sheet_name}»...")
        rows = _read_sheet_rows(wb[sheet_name])
        result.accounts = _import_accounts(db, rows)
        db.flush()
        logger.info(f"Accounts imported: {result.accounts}")
    else:
        msg = f"Лист «{sheet_name}» не найден"
        logger.warning(msg)
        result.errors.append(msg)

    # Import assets
    sheet_name = SHEET_NAMES["assets"]
    if sheet_name in sheet_names_in_file:
        logger.info(f"Importing assets from «{sheet_name}»...")
        rows = _read_sheet_rows(wb[sheet_name])
        result.assets = _import_assets(db, rows)
        db.flush()
        logger.info(f"Assets imported: {result.assets}")
    else:
        msg = f"Лист «{sheet_name}» не найден"
        logger.warning(msg)
        result.errors.append(msg)

    # Import bonds
    sheet_name = SHEET_NAMES["bonds"]
    if sheet_name in sheet_names_in_file:
        logger.info(f"Importing bonds from «{sheet_name}»...")
        rows = _read_sheet_rows(wb[sheet_name])
        result.bonds = _import_bonds(db, rows)
        db.flush()
        logger.info(f"Bonds imported: {result.bonds}")
    else:
        logger.debug(f"Sheet «{sheet_name}» not found (optional)")

    # Import transactions
    sheet_name = SHEET_NAMES["transactions"]
    if sheet_name in sheet_names_in_file:
        logger.info(f"Importing transactions from «{sheet_name}»...")
        rows = _read_sheet_rows(wb[sheet_name])
        result.transactions = _import_transactions(db, rows)
        db.flush()
        logger.info(f"Transactions imported: {result.transactions}")
    else:
        msg = f"Лист «{sheet_name}» не найден"
        logger.warning(msg)
        result.errors.append(msg)

    # Import valuations
    sheet_name = SHEET_NAMES["valuations"]
    if sheet_name in sheet_names_in_file:
        logger.info(f"Importing valuations from «{sheet_name}»...")
        rows = _read_sheet_rows(wb[sheet_name])
        result.valuations = _import_valuations(db, rows)
        db.flush()
        logger.info(f"Valuations imported: {result.valuations}")
    else:
        logger.debug(f"Sheet «{sheet_name}» not found (optional)")

    db.commit()
    wb.close()

    logger.info("=" * 80)
    logger.info(
        f"Import complete: {result.accounts} accounts, {result.assets} assets, "
        f"{result.bonds} bonds, {result.transactions} transactions, "
        f"{result.valuations} valuations"
    )
    if result.errors:
        logger.warning(f"Errors during import:\n" + "\n".join(f"  - {e}" for e in result.errors))
    logger.info("=" * 80)

    flush_all_handlers(logger)
    return result
