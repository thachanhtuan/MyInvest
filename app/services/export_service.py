from __future__ import annotations

import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers

from app.config import EXPORT_LOG_FILE
from app.schemas import PortfolioSummary
from app.utils.logging_util import flush_all_handlers, setup_file_handler

logger = logging.getLogger(__name__)
setup_file_handler(logger, EXPORT_LOG_FILE)  # Configure once at import time


def _style_header(ws, row: int = 1) -> None:
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def export_balance_excel(summary: PortfolioSummary) -> bytes:
    logger.info("=" * 80)
    logger.info("Starting balance export...")
    logger.info(
        f"Portfolio summary: {len(summary.by_account)} accounts, "
        f"{len(summary.holdings)} holdings, {len(summary.by_asset_class)} asset classes"
    )

    try:
        wb = Workbook()

        # Sheet 1: Сводка
        ws = wb.active
        ws.title = "Сводка"
        ws.append(["Показатель", "Значение"])
        _style_header(ws)
        ws.append(["Стоимость портфеля", summary.current_value])
        ws.append(["Вложено", summary.total_invested])
        ws.append(["Прибыль/убыток", summary.profit_loss])
        ws.append(["Доходность, %", summary.profit_loss_pct])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        for row in ws.iter_rows(min_row=2, max_row=5, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '#,##0.00'
        logger.debug("Sheet 'Сводка' created")

        # Sheet 2: По счетам
        ws2 = wb.create_sheet("По счетам")
        ws2.append(["Счёт", "Брокер", "Тип", "Вложено", "Стоимость", "Прибыль", "Доходность %", "Доля %"])
        _style_header(ws2)
        for a in summary.by_account:
            ws2.append([
                a.account_name, a.broker, a.account_type,
                a.total_invested, a.current_value, a.profit_loss,
                a.profit_loss_pct, a.share_pct,
            ])
        for col in ["D", "E", "F"]:
            ws2.column_dimensions[col].width = 16
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 18
        logger.debug(f"Sheet 'По счетам' created with {len(summary.by_account)} accounts")

        # Sheet 3: Позиции
        ws3 = wb.create_sheet("Позиции")
        ws3.append([
            "Тикер", "Название", "Счёт", "Брокер", "Класс актива",
            "Валюта", "Кол-во", "Ср. цена", "Вложено", "Стоимость",
            "Прибыль", "Доходность %",
        ])
        _style_header(ws3)
        for h in summary.holdings:
            ws3.append([
                h.ticker, h.asset_name, h.account_name, h.broker,
                h.asset_class, h.currency, h.quantity, h.avg_price,
                h.total_invested, h.current_value, h.profit_loss,
                h.profit_loss_pct,
            ])
        for col in ["H", "I", "J", "K"]:
            ws3.column_dimensions[col].width = 16
        ws3.column_dimensions["A"].width = 16
        ws3.column_dimensions["B"].width = 25
        logger.debug(f"Sheet 'Позиции' created with {len(summary.holdings)} holdings")

        # Sheet 4: По классам активов
        ws4 = wb.create_sheet("По классам активов")
        ws4.append(["Класс актива", "Стоимость", "Доля %"])
        _style_header(ws4)
        for ac in summary.by_asset_class:
            ws4.append([ac.label, ac.current_value, ac.share_pct])
        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["B"].width = 16
        logger.debug(f"Sheet 'По классам активов' created with {len(summary.by_asset_class)} classes")

        buf = io.BytesIO()
        wb.save(buf)
        file_size = len(buf.getvalue())

        logger.info(f"Balance export completed: {file_size} bytes")
        logger.info("=" * 80)
        flush_all_handlers(logger)

        return buf.getvalue()

    except Exception as e:
        logger.error(f"Export failed: {type(e).__name__}: {e}")
        logger.info("=" * 80)
        flush_all_handlers(logger)
        raise
