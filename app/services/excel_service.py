from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models import (
    Account,
    Asset,
    AssetValuation,
    BondInfo,
    CouponPayment,
    Quote,
    TargetAllocation,
    Transaction,
)


# ─── helpers ──────────────────────────────────────────────────────────────────

def _to_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _sheet_rows(ws) -> List[Dict[str, Any]]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


# ─── import ───────────────────────────────────────────────────────────────────

def import_from_excel(file_like: io.BytesIO, db: Session) -> List[str]:
    wb = load_workbook(file_like, data_only=True)
    log: List[str] = []

    sheet_map = {
        "Счета": _import_accounts,
        "Активы": _import_assets,
        "Облигации": _import_bond_info,
        "Транзакции": _import_transactions,
        "Котировки": _import_quotes,
        "Купоны": _import_coupons,
        "Стоимость актива": _import_valuations,
        "Целевая аллокация": _import_target_allocations,
    }

    for sheet_name, handler in sheet_map.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = handler(ws, db)
            log.append(f"{sheet_name}: {count} rows imported")
        else:
            log.append(f"{sheet_name}: sheet not found, skipped")

    db.commit()
    return log


def _import_accounts(ws, db: Session) -> int:
    count = 0
    for row in _sheet_rows(ws):
        acc_id = _str(row.get("id") or row.get("ID") or row.get("Идентификатор"))
        if not acc_id:
            continue
        obj = db.query(Account).filter(Account.id == acc_id).first()
        data = {
            "id": acc_id,
            "name": _str(row.get("name") or row.get("Название")) or acc_id,
            "type": _str(row.get("type") or row.get("Тип")) or "broker",
            "broker_or_bank": _str(row.get("broker_or_bank") or row.get("Брокер/Банк")) or "",
            "currency": _str(row.get("currency") or row.get("Валюта")),
            "notes": _str(row.get("notes") or row.get("Примечание")),
        }
        if obj:
            for k, v in data.items():
                setattr(obj, k, v)
        else:
            db.add(Account(**data))
        count += 1
    return count


def _import_assets(ws, db: Session) -> int:
    count = 0
    for row in _sheet_rows(ws):
        ticker = _str(row.get("ticker") or row.get("Тикер"))
        if not ticker:
            continue
        obj = db.query(Asset).filter(Asset.ticker == ticker).first()
        data = {
            "ticker": ticker,
            "name": _str(row.get("name") or row.get("Название")) or ticker,
            "asset_class": _str(row.get("asset_class") or row.get("Класс актива")) or "other",
            "currency": _str(row.get("currency") or row.get("Валюта")),
            "isin": _str(row.get("isin") or row.get("ISIN")),
            "notes": _str(row.get("notes") or row.get("Примечание")),
            "target_min": _float(row.get("target_min") or row.get("Мин цель")),
            "target_max": _float(row.get("target_max") or row.get("Макс цель")),
            "target_pct": _float(row.get("target_pct") or row.get("Цель %")),
            "target_date": _to_date(row.get("target_date") or row.get("Целевая дата")),
        }
        if obj:
            for k, v in data.items():
                setattr(obj, k, v)
        else:
            db.add(Asset(**data))
        count += 1
    return count


def _import_bond_info(ws, db: Session) -> int:
    count = 0
    for row in _sheet_rows(ws):
        ticker = _str(row.get("ticker") or row.get("Тикер"))
        if not ticker:
            continue
        obj = db.query(BondInfo).filter(BondInfo.ticker == ticker).first()
        data = {
            "ticker": ticker,
            "face_value": _float(row.get("face_value") or row.get("Номинал")) or 1000.0,
            "base_currency": _str(row.get("base_currency") or row.get("Базовая валюта")),
            "coupon_rate": _float(row.get("coupon_rate") or row.get("Ставка купона")),
            "coupon_sum": _float(row.get("coupon_sum") or row.get("Сумма купона")),
            "coupon_currency": _str(row.get("coupon_currency") or row.get("Валюта купона")),
            "coupon_frequency_year": _int(row.get("coupon_frequency_year") or row.get("Купонов в год")) or 2,
            "coupon_frequency_day": _int(row.get("coupon_frequency_day") or row.get("Период дней")) or 180,
            "maturity_date": _to_date(row.get("maturity_date") or row.get("Дата погашения")),
            "offer_date": _to_date(row.get("offer_date") or row.get("Дата оферты")),
            "first_coupon_date": _to_date(row.get("first_coupon_date") or row.get("Дата первого купона")),
            "is_amortizing": bool(row.get("is_amortizing") or row.get("Амортизация")),
        }
        if obj:
            for k, v in data.items():
                setattr(obj, k, v)
        else:
            db.add(BondInfo(**data))
        count += 1
    return count


def _import_transactions(ws, db: Session) -> int:
    count = 0
    for row in _sheet_rows(ws):
        tx_date = _to_date(row.get("date") or row.get("Дата"))
        ticker = _str(row.get("ticker") or row.get("Тикер"))
        account_id = _str(row.get("account_id") or row.get("Счёт"))
        tx_type = _str(row.get("tx_type") or row.get("Тип"))
        if not all([tx_date, ticker, account_id, tx_type]):
            continue
        obj = Transaction(
            date=tx_date,
            account_id=account_id,
            ticker=ticker,
            tx_type=tx_type,
            quantity=_float(row.get("quantity") or row.get("Количество")),
            price=_float(row.get("price") or row.get("Цена")),
            amount=_float(row.get("amount") or row.get("Сумма")),
            nominal_amount=_float(row.get("nominal_amount") or row.get("Номинальная сумма")),
            nominal_currency=_str(row.get("nominal_currency") or row.get("Номинальная валюта")),
            nkd=_float(row.get("nkd") or row.get("НКД")),
            commission=_float(row.get("commission") or row.get("Комиссия")),
            total_amount=_float(row.get("total_amount") or row.get("Итого")),
            currency=_str(row.get("currency") or row.get("Валюта")),
            broker=_str(row.get("broker") or row.get("Брокер")),
            notes=_str(row.get("notes") or row.get("Примечание")),
        )
        db.add(obj)
        count += 1
    return count


def _import_quotes(ws, db: Session) -> int:
    count = 0
    for row in _sheet_rows(ws):
        ticker = _str(row.get("ticker") or row.get("Тикер"))
        q_date = _to_date(row.get("date") or row.get("Дата"))
        close_price = _float(row.get("close_price") or row.get("Цена закрытия"))
        if not all([ticker, q_date, close_price is not None]):
            continue
        existing = (
            db.query(Quote).filter(Quote.ticker == ticker, Quote.date == q_date).first()
        )
        if existing:
            existing.close_price = close_price
            existing.currency = _str(row.get("currency") or row.get("Валюта"))
            existing.source = _str(row.get("source") or row.get("Источник"))
        else:
            db.add(
                Quote(
                    ticker=ticker,
                    date=q_date,
                    close_price=close_price,
                    currency=_str(row.get("currency") or row.get("Валюта")),
                    source=_str(row.get("source") or row.get("Источник")),
                )
            )
        count += 1
    return count


def _import_coupons(ws, db: Session) -> int:
    count = 0
    for row in _sheet_rows(ws):
        ticker = _str(row.get("ticker") or row.get("Тикер"))
        p_date = _to_date(row.get("payment_date") or row.get("Дата выплаты"))
        amount = _float(row.get("coupon_amount") or row.get("Сумма купона"))
        if not all([ticker, p_date, amount is not None]):
            continue
        db.add(
            CouponPayment(
                ticker=ticker,
                payment_date=p_date,
                coupon_amount=amount,
                currency=_str(row.get("currency") or row.get("Валюта")),
            )
        )
        count += 1
    return count


def _import_valuations(ws, db: Session) -> int:
    count = 0
    for row in _sheet_rows(ws):
        ticker = _str(row.get("ticker") or row.get("Тикер"))
        v_date = _to_date(row.get("date") or row.get("Дата"))
        if not all([ticker, v_date]):
            continue
        db.add(
            AssetValuation(
                date=v_date,
                ticker=ticker,
                value=_float(row.get("value") or row.get("Стоимость")),
                currency=_str(row.get("currency") or row.get("Валюта")),
                nominal_value=_float(row.get("nominal_value") or row.get("Номинальная стоимость")),
                nominal_currency=_str(row.get("nominal_currency") or row.get("Ном. валюта")),
                value_pct=_float(row.get("value_pct") or row.get("% от номинала")),
            )
        )
        count += 1
    return count


def _import_target_allocations(ws, db: Session) -> int:
    count = 0
    for row in _sheet_rows(ws):
        asset_class = _str(row.get("asset_class") or row.get("Класс актива"))
        target_pct = _float(row.get("target_pct") or row.get("Целевой %"))
        if not all([asset_class, target_pct is not None]):
            continue
        existing = (
            db.query(TargetAllocation)
            .filter(TargetAllocation.asset_class == asset_class)
            .first()
        )
        if existing:
            existing.target_pct = target_pct
        else:
            db.add(TargetAllocation(asset_class=asset_class, target_pct=target_pct))
        count += 1
    return count


# ─── export ───────────────────────────────────────────────────────────────────

def export_to_excel(db: Session) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _export_sheet(
        wb,
        "Счета",
        ["id", "name", "type", "broker_or_bank", "currency", "notes"],
        db.query(Account).all(),
    )
    _export_sheet(
        wb,
        "Активы",
        ["ticker", "name", "asset_class", "currency", "isin", "notes",
         "target_min", "target_max", "target_pct", "target_date"],
        db.query(Asset).all(),
    )
    _export_sheet(
        wb,
        "Облигации",
        ["ticker", "face_value", "base_currency", "coupon_rate", "coupon_sum",
         "coupon_currency", "coupon_frequency_year", "coupon_frequency_day",
         "maturity_date", "offer_date", "first_coupon_date", "is_amortizing"],
        db.query(BondInfo).all(),
    )
    _export_sheet(
        wb,
        "Транзакции",
        ["id", "date", "account_id", "ticker", "tx_type", "quantity", "price",
         "amount", "nominal_amount", "nominal_currency", "nkd", "commission",
         "total_amount", "currency", "broker", "notes"],
        db.query(Transaction).order_by(Transaction.date).all(),
    )
    _export_sheet(
        wb,
        "Котировки",
        ["id", "ticker", "date", "close_price", "currency", "source"],
        db.query(Quote).order_by(Quote.date).all(),
    )
    _export_sheet(
        wb,
        "Купоны",
        ["id", "ticker", "payment_date", "coupon_amount", "currency"],
        db.query(CouponPayment).order_by(CouponPayment.payment_date).all(),
    )
    _export_sheet(
        wb,
        "Стоимость актива",
        ["id", "date", "ticker", "value", "currency", "nominal_value", "nominal_currency", "value_pct"],
        db.query(AssetValuation).order_by(AssetValuation.date).all(),
    )
    _export_sheet(
        wb,
        "Целевая аллокация",
        ["id", "asset_class", "target_pct"],
        db.query(TargetAllocation).all(),
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _export_sheet(wb: Workbook, title: str, columns: List[str], records) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(columns)
    for rec in records:
        row = []
        for col in columns:
            val = getattr(rec, col, None)
            if isinstance(val, date):
                val = val.isoformat()
            row.append(val)
        ws.append(row)
